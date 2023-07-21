import numpy as np
import pandas as pd
import datetime
import openpyxl
import matplotlib.pyplot as plt
import os
import math as math
import itertools
import pickle
import seaborn as sns
import tensorflow as tf

from sklearn.metrics import confusion_matrix
from sklearn.metrics import roc_curve
from sklearn.metrics import auc

import torch

def plot_training(history, plot_training_dir):
    if not isinstance(history, pd.DataFrame):
        history = pd.DataFrame.from_dict(history, orient='index').transpose()

    if 'train_loss' in history.columns and 'val_loss' in history.columns:
        plt.figure(figsize=(8, 6))
        for c, color, in zip(['train_loss', 'val_loss'],['r', 'b']):
            plt.plot(history[c], label=c, color=color)
        plt.title("Training and validation loss")
        plt.xlabel('Epochs')
        plt.ylabel('Average Negative Log Likelihood')
        plt.legend()
        plt.savefig(os.path.join(plot_training_dir, "train_val_loss.png"), dpi=400, format='png')
        plt.show()
    if 'train_acc' in history.columns and 'val_acc' in history.columns:
        plt.figure(figsize=(8, 6))
        for c, color, in zip(['train_acc', 'val_acc'],['r', 'b']):
            plt.plot(history[c], label=c, color=color)
        plt.title('Training and Validation Accuracy')
        plt.xlabel('Epochs')
        plt.ylabel('Average Accuracy')
        plt.ylim([-0.1, 1.1])
        plt.legend()
        plt.savefig(os.path.join(plot_training_dir, "train_val_acc.png"), dpi=400, format='png')
        plt.show()


def write_readme(reports_dir, tag_name, filename, **attributes):
    """
    Function to write on .txt file a dictionary
    Args:
        reports_dir (String): folder to save data
        tag_name (String): header to include in .txt file
        filename
        attributes: attributes of dictionary
    """
    print('\n[INFO]', tag_name, sep=' ')
    [print(name, value, sep=': ') for name, value in attributes.items()]
    with open(os.path.join(reports_dir, filename), "a") as file:
        if tag_name[:] == '/nStart':
            file.write("START: " + datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S"))
        file.write(tag_name)
        [file.write("\n{}: {}".format(name, value)) for name, value in attributes.items()]
        file.write("\n\n")

def append_df_to_excel(filename, dataframe, sheet_name='Sheet1', startrow=None,  **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename] into [sheet_name] Sheet. If [filename] doesn't exist, then this function will create it.
        Parameters:
            filename: File path or existing ExcelWriter (Example: '/path/to/file.xlsx')
            dataframe: dataframe to save to workbook
            sheet_name: Name of sheet which will contain DataFrame (default: 'Sheet1')
            startrow: upper left cell row to dump data frame. Per default (startrow=None) calculate the last row in the existing DF and write to the next row
            to_excel_kwargs: arguments which will be passed to `DataFrame.to_excel()`  [can be dictionary]
        Returns:
             None
    """
    header = True
    if 'engine' in to_excel_kwargs: # ignore [engine] parameter if it was passed
        to_excel_kwargs.pop('engine')
    writer = pd.ExcelWriter(filename, engine='openpyxl')
    try:
        writer.book = openpyxl.load_workbook(filename)  # try to open an existing workbook
        if startrow is None and sheet_name in writer.book.sheetnames:  # get the last row in the existing Excel sheet if it was not specified explicitly
            startrow = writer.book[sheet_name].max_row
            header = False
        writer.sheets = {ws.title: ws for ws in writer.book.worksheets}  # copy existing sheets
    except FileNotFoundError:  # file does not exist yet, we will create it
        pass
    if startrow is None:
        startrow = 0
    dataframe.to_excel(writer, sheet_name, startrow=startrow, header=header, **to_excel_kwargs)  # write out the new sheet
    writer.save()  # save the workbook


def plot_acc(outdir, acc, info='acc', channel=3):

    if acc.shape[0] == channel:
        acc = torch.transpose(acc, dim0=0, dim1=1).detach().cpu().numpy()
    if acc.shape[0] == 1:
        # from Tensofrlow to Numpy
        acc = acc.numpy()[0]

    time = np.arange(acc.shape[0])
    fig, (ax1, ax2, ax3) = plt.subplots(3, 1)

    ax1.plot(time, acc[:, 0], 'black', linewidth=0.7)
    ax1.axis(xmin=0, xmax=2490)
    ax1.set_ylabel('x')

    ax2.plot(time, acc[:, 1], 'black', linewidth=0.7)
    ax2.axis(xmin=0, xmax=2490)
    ax2.set_ylabel('y')

    ax3.plot(time, acc[:, 2], 'black', linewidth=0.7)
    ax3.axis(xmin=0, xmax=2490)
    ax3.set_ylabel('z')

    plt.tight_layout()
    plt.savefig(os.path.join(outdir, f"{info}.png"), dpi=400, format='png')
    plt.show()

def plot_vel(outdir, vel, info='vel'):

    if vel.shape[0] == 1:
        vel = vel.numpy()[0]

    time = np.arange(vel.shape[0])
    # One axis plot.
    fig, ax = plt.subplots(1, 1)
    ax.plot(time, vel, 'black', linewidth=1.0)
    ax.axis(xmin=0, xmax=41)
    ax.set_ylabel('pos')

    plt.tight_layout()
    plt.savefig(os.path.join(outdir, f"{info}.png"), dpi=400, format='png')
    plt.show()


# gradcam/ig
def plot_heatmap(outdir, acc, heatmap, info):

    if acc.shape[0] == 1:
        acc = acc.numpy()[0]

    heatmap_x = np.expand_dims(np.array(heatmap[:,0]), axis=1)
    heatmap_y = np.expand_dims(np.array(heatmap[:,1]), axis=1)
    heatmap_z = np.expand_dims(np.array(heatmap[:,2]), axis=1)

    # Normalization between 0 and 1
    acc_min, acc_max = np.min(acc), np.max(acc)
    acc_norm = (acc - acc_min) / (acc_max - acc_min)
    acc_norm_x = acc_norm[:,0]
    acc_norm_y = acc_norm[:,1]
    acc_norm_z = acc_norm[:,2]

    # Overlap of the signal with the heatmap data from GradCam
    print(f'Overlap of acceleration signal and heatmap')
    time = np.arange(acc_norm_x.shape[0])
    fig, (ax1, ax2, ax3) = plt.subplots(3, 1, figsize = (8,4))

    ax1.plot(time, acc_norm_x, color = 'black', linewidth=1)
    ax_h1 = sns.heatmap(
        heatmap_x.transpose(), ax = ax1, cmap = 'Reds',  xticklabels = False, yticklabels = False, vmin = 0, vmax = 1, cbar= False)
    ax1.axis(ymin=0, ymax=1)
    ax1.set_ylabel('x')

    ax2.plot(time, acc_norm_y, color = 'black', linewidth=1)
    ax_h2 = sns.heatmap(
        heatmap_y.transpose(),  ax=ax2, cmap = 'Reds', xticklabels = False,  yticklabels = False, vmin = 0, vmax = 1,cbar= False)
    ax2.axis(ymin=0, ymax=1)
    ax2.set_ylabel('y')

    ax3.plot(time, acc_norm_z, color = 'black', linewidth=1)
    ax_h3 = sns.heatmap(
        heatmap_z.transpose(), ax = ax3, cmap = 'Reds',  xticklabels = 500, yticklabels = False, vmin = 0, vmax = 1, cbar = False)
    ax3.axis(ymin=0, ymax=1)
    ax3.set_ylabel('z')
    plt.tight_layout()
    plt.show()
    fig.savefig(os.path.join(outdir, f'{info}_xai.png'), format='png', dpi=400)

def plot_mts_heatmap(outdir, acc, heatmap, info,channel=3):
    if acc.shape[0] == channel:
        acc = torch.transpose(acc, dim0=0, dim1=1).detach().cpu().numpy()
    if acc.shape[0] == 1:
        # from Tensofrlow to Numpy
        acc = acc.numpy()[0]

    # creation of a tensor of 2490 elements that goes from 1 to 2490
    time = np.arange(acc.shape[0])
    fig, (ax1, ax2, ax3, ax4) = plt.subplots(4, 1, figsize=(10, 8))

    ax1.plot(time, acc[:, 0], 'black', linewidth=0.7)
    ax1.axis(xmin=0, xmax=2490)
    ax1.set_ylabel('x')

    ax2.plot(time, acc[:, 1], 'black', linewidth=0.7)
    ax2.axis(xmin=0, xmax=2490)
    ax2.set_ylabel('y')

    ax3.plot(time, acc[:, 2], 'black', linewidth=0.7)
    ax3.axis(xmin=0, xmax=2490)
    ax3.set_ylabel('z')

    ax4.plot(time, heatmap, 'red', linewidth=1)
    ax4.axis(xmin=0, xmax=2490)
    ax4.set_ylabel('heatmap')

    plt.tight_layout()
    fig.savefig(os.path.join(outdir, f'{info}.png'), format='png', dpi=400)
    plt.show()

# lime
def heatmap_lime(outdir, acc, mask, info='heatmap_lime'):

    if acc.shape[0] == 1:
        # from Tensofrlow to Numpy
        acc = acc.numpy()[0]

    time = np.arange(acc.shape[0])

    fig, (ax1, ax2, ax3) = plt.subplots(3, 1, figsize=(8, 4))
    ax1.plot(time, acc[:, 0], color='black', linewidth=1)
    ax1.fill_between(time, np.amin(acc[:, 0]), np.amax(acc[:, 0]), where=mask[:, 0] == 1, facecolor='green',  alpha=0.5)
    ax1.set_ylabel('x')
    ax1.axis(xmin=0, xmax=2490)
    ax1.get_xaxis().set_ticks([])

    ax2.plot(time, acc[:, 1], color='black', linewidth=1)
    ax2.fill_between(time, np.amin(acc[:, 1]), np.amax(acc[:, 1]), where=mask[:, 1] == 1, facecolor='green',    alpha=0.5)
    ax2.set_ylabel('y')
    ax2.axis(xmin=0, xmax=2490)
    ax2.get_xaxis().set_ticks([])

    ax3.plot(time, acc[:, 2], color='black', linewidth=1)
    ax3.fill_between(time, np.amin(acc[:, 2]), np.amax(acc[:, 2]), where=mask[:, 2] == 1, facecolor='green',   alpha=0.5)
    ax3.set_ylabel('z')
    ax3.axis(xmin=0, xmax=2490)

    plt.tight_layout()
    fig.savefig(os.path.join(outdir, f'{info}.png'), format='png', dpi=400)
    plt.show()