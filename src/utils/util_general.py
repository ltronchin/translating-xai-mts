import argparse
from pathlib import Path
import os
import numpy as np
import random
import openpyxl
import torch
import requests

def notification_ifttt(info):
    private_key = "isnY23hWBGyL-mF7F18BUAC-bGAN6dx1UAPoqnfntUa"
    url = "https://maker.ifttt.com/trigger/Notification/json/with/key/" + private_key
    requests.post(url, data={'value1': "Update --- "+ str(info)})

def get_args_cae_generali():
    parser = argparse.ArgumentParser(description="Configuration File")
    parser.add_argument("-f", "--cfg_file", help="Path of Configuration File", type=str)
    parser.add_argument("-d", "--device_type", help="Select CPU or GPU", type=str, default="gpu")
    parser.add_argument("-g", "--gpu", help="CUDA device", type=str, default="cuda:0")
    parser.add_argument("-i", "--id_run", help="id run", type=int, default=0)
    parser.add_argument("--mode", default='client')
    parser.add_argument("--port", default=53667)
    return parser.parse_args()

def rgb2gray(rgb):
    return np.dot(rgb[..., :3], [0.2989, 0.5870, 0.1140])

def seed_all(seed): # for deterministic behaviour
    if not seed:
        seed = 42
    print("Using Seed : ", seed)

    os.environ['PYTHONHASHSEED'] = str(seed)
    torch.cuda.empty_cache()
    torch.manual_seed(seed) # Set torch pseudo-random generator at a fixed value
    torch.cuda.manual_seed_all(seed)
    torch.cuda.manual_seed(seed)
    np.random.seed(seed) # Set numpy pseudo-random generator at a fixed value
    random.seed(seed) # Set python built-in pseudo-random generator at a fixed value
    torch.backends.cudnn.deterministic = True
    torch.backends.cudnn.benchmark = False

def create_dir(dir): # function to create directory
    if not os.path.exists(dir):
        Path(dir).mkdir(parents=True, exist_ok=True) # with parents 'True' creates all tree/nested folder

def create_path(*path_list, f=None):
    f = path_list[0]
    for i in range(1, len(path_list)):
        path = str(path_list[i])
        f = os.path.join(f, path)
    return f

def delete_file(file_path):
    try:
        os.remove(file_path)
    except FileNotFoundError:
        pass

def write_excel(*keys, dictionary, sheet_idx, col_position=1, wb=None, ws=None, default='not found'):
    """ Function to save the patients data splits (ids and labels) to file excel (openpyxl library is used)
    Args:
        keys: string values representing the field from the dictionary to save in excel
        dictionary: dictionary representing the data
        sheet_idx: int value representing the index of the current worksheet
        col_position: int value representing the coordinate of the column to begin to start the writing
        wb: existing workbook
        ws: existing worksheet
        default: default values to return if the inserted key do not exist in dictionary
    Returns:
        wb: workbook
        ws: worksheet
    """

    if wb is None:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '_fold' + str(sheet_idx)
    if ws is None:
        ws = wb.create_sheet(title='_fold' + str(sheet_idx))

    ws.cell(row=1, column=col_position, value=keys[0].split('_')[1] + '_id')
    ws.cell(row=1, column=col_position + 1, value=keys[0].split('_')[1] + '_label')

    for idx in range(dictionary.get(keys[0], default).shape[0]):
        [ws.cell(row=idx + 2, column=col_position + column, value=dictionary.get(key, default)[idx, 0]) for key, column in zip(keys, np.arange(len(keys)))]
    return wb, ws